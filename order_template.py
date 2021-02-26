from order_template_gui import *
from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
import sys
from PyQt5.QtWidgets import QFileDialog, QTableWidgetItem
from datetime import *
import os
from xlrd import open_workbook


big_dict = {}
stock_dict = {}
stock_path = ''
stock_start_row = 6
stock_ysku_col = 2
stock_sku_col = 1
stock_value_col = 6
assr_dict = {}
assortment_path = ''
assr_ysku_col = 29
assr_sku_col = 2
assr_art_col = 10
stat_dict = {}
stat_sales_path = ''
stat_start_row = 3
stat_sku_col = 4
stat_value_col = 6
route_path = ''
route_dict = {}
kotelniki_list = []
sofino_list = []
ex_path_1 = ''
ex_path_2 = ''
ex_path_3 = ''
ex_path_4 = ''
ex_set = set()


class MyWin(QtWidgets.QMainWindow):
    def __init__(self, parent=None):
        QtWidgets.QWidget.__init__(self, parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        # кнопки
        self.ui.stock_Button.clicked.connect(lambda: self.ui.stock_path.setText(QFileDialog.getOpenFileName()[0]))
        self.ui.assortment_Button.clicked.connect(
            lambda: self.ui.assortment_path.setText(QFileDialog.getOpenFileName()[0]))
        self.ui.stat_sales_Button.clicked.connect(
            lambda: self.ui.stat_sales_path.setText(QFileDialog.getOpenFileName()[0]))
        self.ui.route_Button.clicked.connect(
            lambda: self.ui.route_kotelniki_path.setText(QFileDialog.getOpenFileName()[0]))
        self.ui.create_Button.clicked.connect(self.get_paths)
        self.ui.ex_Button_1.clicked.connect(lambda: self.ui.ex_path_1.setText(QFileDialog.getOpenFileName()[0]))
        self.ui.ex_Button_2.clicked.connect(lambda: self.ui.ex_path_2.setText(QFileDialog.getOpenFileName()[0]))
        self.ui.ex_Button_3.clicked.connect(lambda: self.ui.ex_path_3.setText(QFileDialog.getOpenFileName()[0]))
        self.ui.ex_Button_4.clicked.connect(lambda: self.ui.ex_path_4.setText(QFileDialog.getOpenFileName()[0]))

    def get_paths(self):
        global stock_path, assortment_path, stat_sales_path, route_path, ex_path_1, ex_path_2, ex_path_3, ex_path_4
        stock_path = self.ui.stock_path.text()
        assortment_path = self.ui.assortment_path.text()
        stat_sales_path = self.ui.stat_sales_path.text()
        route_path = self.ui.route_kotelniki_path.text()
        ex_path_1 = self.ui.ex_path_1.text()
        ex_path_2 = self.ui.ex_path_2.text()
        ex_path_3 = self.ui.ex_path_3.text()
        ex_path_4 = self.ui.ex_path_4.text()

        self.check_ex()
        self.create_assortment_dict()
        self.create_stock_dict()
        self.create_stat_dict()
        self.create_big_dict()
        self.create_routes()
        self.write_files()


    def check_ex(self):
        if ex_path_1:
            self.create_ex(ex_path_1)
        if ex_path_2:
            self.create_ex(ex_path_2)
        if ex_path_3:
            self.create_ex(ex_path_3)
        if ex_path_4:
            self.create_ex(ex_path_4)

    def create_ex(self, path):
        try:
            wb = load_workbook(path)
            ws = wb.worksheets[0]
            mr = ws.max_row
            for i in range(1, mr + 1):
                ex_set.add(str(ws.cell(row=i, column=1).value).strip())
        except:
            wb = open_workbook(path, on_demand=True)
            ws = wb.sheet_by_index(0)
            mr = ws.nrows
            for i in range(0, mr):
                ex_set.add(str(ws.cell_value(rowx=i, colx=0)).strip())

    def write_files(self):
        # определяем стили
        fill = PatternFill(fill_type='solid',
                           start_color='c1c1c1',
                           end_color='c2c2c2')

        today = datetime.today()
        today = today.strftime('%d.%m.%Y')
        file_name = today + '_Order.xlsx'
        wb = Workbook()
        ws_k = wb.active
        ws_k.title = 'Котельники'

        for row in kotelniki_list:
            ws_k.append(row)
        # раскрашивание фона для заголовков
        ws_k['A1'].fill = fill
        ws_k['B1'].fill = fill
        ws_k['C1'].fill = fill
        ws_k['D1'].fill = fill
        ws_k['E1'].fill = fill

        ws_s = wb.create_sheet('Софьино')

        for row in sofino_list:
            ws_s.append(row)
            # раскрашивание фона для заголовков
        ws_s['A1'].fill = fill
        ws_s['B1'].fill = fill
        ws_s['C1'].fill = fill
        ws_s['D1'].fill = fill
        ws_s['E1'].fill = fill

        wb.save(file_name)
        os.startfile(file_name)

    def create_routes(self):
        global route_dict, kotelniki_list, sofino_list
        kotelniki_list.append(['SKU', 'YSKU', 'ART', 'Stock', 'Order'])
        sofino_list.append(['SKU', 'YSKU', 'ART', 'Stock', 'Order'])
        route = load_workbook(route_path)
        route_sheet = route.worksheets[0]
        route_dict = {x[0].value for x in route_sheet}

        for big_key in big_dict:
            if big_key in route_dict:
                kotelniki_list.append([big_key, big_dict[big_key]['YSKU'], big_dict[big_key]['ART'],
                                       big_dict[big_key]['Stock'], big_dict[big_key]['Order']])
            else:
                sofino_list.append([big_key, big_dict[big_key]['YSKU'], big_dict[big_key]['ART'],
                                       big_dict[big_key]['Stock'], big_dict[big_key]['Order']])

    def create_big_dict(self):
        to_del = []
        for big_key, big_val in big_dict.items():
            if big_key in ex_set:
                to_del.append(big_key)
                continue
            if big_key in stat_dict:
                if big_val['Stock'] > stat_dict[big_key]['Order'] / 2:
                    to_del.append(big_key)
                else:
                    big_dict[big_key]['Order'] = stat_dict[big_key]['Order']
            elif big_val['Stock'] > 0:
                to_del.append(big_key)
        for key in to_del:
            big_dict.pop(key)
        for key in big_dict:
            big_dict[key]['ART'] = assr_dict[key]['ART']

    def create_stat_dict(self):
        global stat_dict
        stat = load_workbook(stat_sales_path)
        stat_sheet = stat.worksheets[0]
        mr = stat_sheet.max_row

        for i in range(stat_start_row, mr + 1):
            sku = str(stat_sheet.cell(row=i, column=stat_sku_col).value).strip()
            order = int(stat_sheet.cell(row=i, column=stat_value_col).value)
            if sku in big_dict:
                if sku not in stat_dict:
                    stat_dict[sku] = {'YSKU': '', 'ART': 'УТ', 'Stock': 0, 'Order': order}
                elif sku in stat_dict:
                    stat_dict[sku]['Order'] += order

    def create_assortment_dict(self):
        global assr_dict
        assr = load_workbook(assortment_path)
        assr_sheet = assr.worksheets[2]
        assr_dict = {x[1].value: {'YSKU': x[28].value, 'ART': x[9].value, 'Stock': 0, 'Order': 5}
                     for i, x in enumerate(assr_sheet) if i > 3}

    def create_stock_dict(self):
        global stock_dict, big_dict
        stock = load_workbook(stock_path)
        stock_sheet = stock.worksheets[0]
        mr = stock_sheet.max_row

        for i in range(stock_start_row, mr + 1):
            sku = str(stock_sheet.cell(row=i, column=stock_sku_col).value).strip()
            ysku = str(stock_sheet.cell(row=i, column=stock_ysku_col).value).strip()
            stock_stock = int(stock_sheet.cell(row=i, column=stock_value_col).value)
            if sku in assr_dict:
                if sku not in stock_dict:
                    stock_dict[sku] = {'YSKU': ysku, 'ART': 'УТ', 'Stock': stock_stock, 'Order': 5}
                elif sku in stock_dict:
                    stock_dict[sku]['Stock'] += stock_stock

        big_dict = assr_dict.copy()
        big_dict.update(stock_dict)


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    myapp = MyWin()
    myapp.show()
    sys.exit(app.exec_())
